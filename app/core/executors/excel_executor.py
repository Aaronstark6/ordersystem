from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from app.export.model import ExportStrategy


class ExcelExecutionError(Exception):
    pass


@dataclass
class ExportOperationResult:
    operation_id: str
    operation_type: str
    status: str
    message: str
    target: Dict[str, Any] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ExcelExportResult:
    output_file_path: str
    success_count: int = 0
    skipped_count: int = 0
    failed_count: int = 0
    operation_results: List[ExportOperationResult] = field(default_factory=list)


def _get_target_value(target: dict[str, Any], key: str) -> Any:
    value = target.get(key)
    if value is None or value == "":
        raise ExcelExecutionError(f"导出目标缺少 {key}")
    return value


def _get_worksheet(workbook: Any, sheet_name: str) -> Any:
    if sheet_name not in workbook.sheetnames:
        raise ExcelExecutionError(f"工作表不存在: {sheet_name}")
    return workbook[sheet_name]


def _write_value(workbook: Any, target: Dict[str, Any], value: Any) -> None:
    sheet_name = _get_target_value(target, "sheet_name")
    cell = _get_target_value(target, "cell")
    worksheet = _get_worksheet(workbook, sheet_name)
    worksheet[cell] = value


def _resolve_table_write_cell(
    worksheet: Any,
    row: int,
    column: int,
) -> tuple[Any | None, str | None, bool]:
    target_cell = worksheet.cell(row=row, column=column)
    if not isinstance(target_cell, MergedCell):
        return target_cell, None, False

    for merged_range in worksheet.merged_cells.ranges:
        if target_cell.coordinate not in merged_range:
            continue
        writable_cell = worksheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        )
        return (
            writable_cell,
            (
                f"write_table redirected merged cell {target_cell.coordinate} "
                f"to {writable_cell.coordinate}"
            ),
            True,
        )

    return (
        None,
        f"write_table skipped merged cell {target_cell.coordinate}: range not found",
        True,
    )


def _write_table(workbook: Any, target: Dict[str, Any], value: Any) -> tuple[int, int, List[str]]:
    sheet_name = _get_target_value(target, "sheet_name")
    start_cell = _get_target_value(target, "start_cell")
    if not isinstance(value, list):
        raise ExcelExecutionError("write_table 的 value 必须是 list")

    worksheet = _get_worksheet(workbook, sheet_name)
    column_letters, start_row = coordinate_from_string(start_cell)
    start_column = column_index_from_string(column_letters)
    written_count = 0
    skipped_count = 0
    warnings: List[str] = []
    redirected_targets: set[str] = set()

    for row_offset, row_values in enumerate(value):
        values = row_values if isinstance(row_values, (list, tuple)) else [row_values]
        for column_offset, cell_value in enumerate(values):
            writable_cell, warning, redirected = _resolve_table_write_cell(
                worksheet,
                row=start_row + row_offset,
                column=start_column + column_offset,
            )
            if warning:
                warnings.append(warning)
            if writable_cell is None:
                skipped_count += 1
                continue

            if redirected and writable_cell.coordinate in redirected_targets:
                skipped_count += 1
                warnings.append(
                    f"write_table skipped duplicate merged target {writable_cell.coordinate}"
                )
                continue

            writable_cell.value = cell_value
            written_count += 1
            redirected_targets.add(writable_cell.coordinate)

    return written_count, skipped_count, warnings


def _validate_insert_image(
    workbook: Any,
    target: Dict[str, Any],
    value: Any,
) -> None:
    sheet_name = _get_target_value(target, "sheet_name")
    _get_target_value(target, "cell")
    _get_worksheet(workbook, sheet_name)
    if value is None or value == "":
        raise ExcelExecutionError("insert_image 缺少图片值")


def _write_choice_mark(
    workbook: Any,
    coordinate: Dict[str, Any],
    fallback_target: Dict[str, Any],
) -> None:
    sheet_name = coordinate.get("sheet_name") or fallback_target.get("sheet_name")
    if not sheet_name:
        raise ExcelExecutionError("选择项坐标缺少 sheet_name")

    worksheet = _get_worksheet(workbook, sheet_name)
    cell = coordinate.get("cell")
    if cell:
        worksheet[cell] = "✓"
        return

    row = coordinate.get("row")
    column = coordinate.get("column")
    if row is None or column is None:
        raise ExcelExecutionError("选择项坐标缺少 cell 或 row/column")

    worksheet.cell(row=row, column=column, value="✓")


def _apply_set_choice_operation(
    workbook: Any,
    target: Dict[str, Any],
    value: Any,
) -> tuple[int, List[str]]:
    if not isinstance(value, dict):
        raise ExcelExecutionError("set_choice 的 value 必须是 dict")

    choice_mode = value.get("choice_mode")
    final_value = value.get("final_value")

    if choice_mode == "value":
        _write_value(workbook, target, final_value)
        return 1, []

    if choice_mode == "dropdown":
        if target and final_value is not None and final_value != "":
            _write_value(workbook, target, final_value)
            return 1, []
        return 0, ["dropdown 缺少可写 target 或 final_value"]

    if choice_mode not in {
        "checkbox_group",
        "radio_group",
        "multiselect",
    }:
        return 0, [f"暂不支持 choice_mode: {choice_mode}"]

    final_selected_values = value.get("final_selected_values", [])
    option_details = value.get("option_details", [])
    if not isinstance(final_selected_values, list):
        raise ExcelExecutionError("set_choice 的 final_selected_values 必须是 list")
    if not isinstance(option_details, list):
        raise ExcelExecutionError("set_choice 的 option_details 必须是 list")

    write_count = 0
    warnings: List[str] = []
    for option in option_details:
        if not isinstance(option, dict):
            warnings.append("跳过无法解析的选择项")
            continue

        option_key = option.get("option_key")
        option_value = option.get("value")
        is_selected = any(
            selected == option_key or selected == option_value
            for selected in final_selected_values
        )
        if not is_selected:
            # Preserve unselected template content in the first implementation.
            continue

        coordinate = option.get("coordinate")
        if not isinstance(coordinate, dict):
            warnings.append(f"选择项缺少可用坐标: {option_key or option_value}")
            continue

        try:
            _write_choice_mark(workbook, coordinate, target)
            write_count += 1
        except ExcelExecutionError as exc:
            warnings.append(f"{option_key or option_value}: {exc}")

    return write_count, warnings


def execute_excel_export(
    template_file_path: str,
    output_file_path: str,
    export_strategy: ExportStrategy,
) -> ExcelExportResult:
    template_path = Path(template_file_path)
    output_path = Path(output_file_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Excel 模板文件不存在: {template_file_path}")

    workbook = load_workbook(filename=template_path)
    result = ExcelExportResult(output_file_path=str(output_path))

    for operation in export_strategy.operations:
        operation_result = ExportOperationResult(
            operation_id=operation.operation_id,
            operation_type=operation.operation_type,
            status="",
            message="",
            target=dict(operation.target),
            metadata=dict(operation.metadata),
        )

        try:
            if operation.operation_type == "write_value":
                _write_value(workbook, operation.target, operation.value)
                operation_result.status = "success"
                operation_result.message = "单元格值写入成功"
                result.success_count += 1
            elif operation.operation_type == "write_table":
                write_count, skipped_count, warnings = _write_table(
                    workbook,
                    operation.target,
                    operation.value,
                )
                if warnings:
                    operation_result.metadata["warnings"] = warnings
                if skipped_count:
                    operation_result.metadata["skipped_cell_count"] = skipped_count
                operation_result.metadata["written_cell_count"] = write_count
                if write_count > 0:
                    operation_result.status = "success"
                    operation_result.message = (
                        f"table values written: {write_count}"
                    )
                    if skipped_count:
                        operation_result.message += (
                            f", skipped cells: {skipped_count}"
                        )
                    result.success_count += 1
                else:
                    operation_result.status = "skipped"
                    operation_result.message = (
                        warnings[0] if warnings else "no writable table cells"
                    )
                    result.skipped_count += 1
            elif operation.operation_type == "insert_image":
                _validate_insert_image(
                    workbook,
                    operation.target,
                    operation.value,
                )
                operation_result.status = "skipped"
                operation_result.message = "图片插入尚未实现，已完成占位校验"
                result.skipped_count += 1
            elif operation.operation_type == "set_choice":
                write_count, warnings = _apply_set_choice_operation(
                    workbook,
                    operation.target,
                    operation.value,
                )
                if warnings:
                    operation_result.metadata["warnings"] = warnings
                if write_count > 0:
                    operation_result.status = "success"
                    operation_result.message = (
                        f"选择写入成功，共写入 {write_count} 项"
                    )
                    if warnings:
                        operation_result.message += (
                            f"，跳过 {len(warnings)} 项"
                        )
                    result.success_count += 1
                else:
                    operation_result.status = "skipped"
                    operation_result.message = (
                        warnings[0] if warnings else "没有可写的选择项"
                    )
                    result.skipped_count += 1
            else:
                operation_result.status = "skipped"
                operation_result.message = (
                    f"暂不支持导出操作: {operation.operation_type}"
                )
                result.skipped_count += 1
        except Exception as exc:
            operation_result.status = "failed"
            operation_result.message = str(exc)
            result.failed_count += 1

        result.operation_results.append(operation_result)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return result
