from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from app.contracts.template_analysis_result import CellInfo, MergedRangeInfo, SheetInfo


def _find_merged_range(cell_row: int, cell_column: int, merged_ranges: List[MergedRangeInfo]) -> Optional[str]:
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= cell_row <= merged_range.max_row
            and merged_range.min_col <= cell_column <= merged_range.max_col
        ):
            return merged_range.range_ref
    return None


def read_excel_template(file_path: str) -> List[SheetInfo]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 模板文件不存在: {file_path}")

    workbook = load_workbook(filename=path, data_only=False)
    sheets: List[SheetInfo] = []

    for worksheet in workbook.worksheets:
        merged_ranges: List[MergedRangeInfo] = []
        for merged in worksheet.merged_cells.ranges:
            merged_ranges.append(
                MergedRangeInfo(
                    sheet_name=worksheet.title,
                    range_ref=str(merged),
                    min_row=merged.min_row,
                    min_col=merged.min_col,
                    max_row=merged.max_row,
                    max_col=merged.max_col,
                )
            )

        row_heights: Dict[int, float] = {}
        for row_index, row_dimension in worksheet.row_dimensions.items():
            if row_dimension.height is not None:
                row_heights[int(row_index)] = float(row_dimension.height)

        column_widths: Dict[str, float] = {}
        for column_key, column_dimension in worksheet.column_dimensions.items():
            if column_dimension.width is not None:
                column_widths[str(column_key)] = float(column_dimension.width)

        cells: List[CellInfo] = []
        for row in worksheet.iter_rows():
            for cell in row:
                value: Any = cell.value
                has_value = value is not None and str(value).strip() != ""
                has_style = bool(cell.has_style)

                if not has_value and not has_style:
                    continue

                merged_range = _find_merged_range(cell.row, cell.column, merged_ranges)

                cells.append(
                    CellInfo(
                        sheet_name=worksheet.title,
                        cell=cell.coordinate,
                        row=cell.row,
                        column=cell.column,
                        value=value,
                        data_type=str(cell.data_type),
                        is_merged=merged_range is not None,
                        merged_range=merged_range,
                        has_style=has_style,
                        number_format=cell.number_format,
                    )
                )

        sheets.append(
            SheetInfo(
                sheet_name=worksheet.title,
                max_row=worksheet.max_row,
                max_column=worksheet.max_column,
                cells=cells,
                merged_ranges=merged_ranges,
                row_heights=row_heights,
                column_widths=column_widths,
            )
        )

    return sheets
