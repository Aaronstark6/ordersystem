from dataclasses import dataclass, field
from typing import Any, Dict, List

from openpyxl.utils import get_column_letter

from app.contracts.template_analysis_result import CellInfo, FieldLabelCandidate, SheetInfo
from app.document_model.coordinates import Coordinate


@dataclass
class FieldCandidate:
    field_key: str
    label: str
    normalized_name: str
    source: str
    confidence: float
    coordinate: Coordinate
    metadata: Dict[str, Any] = field(default_factory=dict)


def build_field_candidate(
    field_key: str,
    label: str,
    normalized_name: str,
    source: str,
    confidence: float,
    coordinate: Coordinate,
    metadata: Dict[str, Any] | None = None,
) -> FieldCandidate:
    return FieldCandidate(
        field_key=field_key,
        label=label,
        normalized_name=normalized_name,
        source=source,
        confidence=confidence,
        coordinate=coordinate,
        metadata=dict(metadata or {}),
    )


_MAX_LABEL_LENGTH = 80

_CHINESE_LABEL_HINTS = (
    "客户",
    "客户名称",
    "公司",
    "公司名称",
    "联系人",
    "电话",
    "手机",
    "邮箱",
    "地址",
    "产品",
    "产品名称",
    "数量",
    "单价",
    "金额",
    "日期",
    "备注",
    "订单号",
)

_ENGLISH_LABEL_HINTS = (
    "customer",
    "customer name",
    "client",
    "company",
    "company name",
    "contact",
    "phone",
    "tel",
    "email",
    "address",
    "product",
    "product name",
    "quantity",
    "qty",
    "unit price",
    "amount",
    "date",
    "remark",
    "notes",
    "order no",
    "order number",
)


def _normalize_label_text(value: object) -> str:
    return str(value).strip().rstrip(":：").strip()


def _normalize_english_text(text: str) -> str:
    return " ".join(text.lower().split())


def _looks_like_label(value: object) -> bool:
    if value is None:
        return False

    raw_text = str(value).strip()
    if not raw_text:
        return False

    label_text = _normalize_label_text(raw_text)
    if not label_text:
        return False
    if len(label_text) > _MAX_LABEL_LENGTH:
        return False
    if label_text.isnumeric():
        return False

    has_colon = raw_text.endswith((":", "："))
    if has_colon:
        return True

    english_text = _normalize_english_text(label_text)
    if english_text in _ENGLISH_LABEL_HINTS:
        return True

    return any(hint in label_text for hint in _CHINESE_LABEL_HINTS)


def _is_blank_target(cell: CellInfo | None) -> bool:
    if cell is None:
        return True
    if cell.value is None:
        return True
    return str(cell.value).strip() == ""


def _cell_ref(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _infer_target_cell(sheet: SheetInfo, cell: CellInfo) -> str | None:
    cell_by_position = {
        (sheet_cell.row, sheet_cell.column): sheet_cell
        for sheet_cell in sheet.cells
    }

    right_column = cell.column + 1
    if right_column <= sheet.max_column:
        right_cell = cell_by_position.get((cell.row, right_column))
        if _is_blank_target(right_cell):
            return _cell_ref(cell.row, right_column)

    below_row = cell.row + 1
    if below_row <= sheet.max_row:
        below_cell = cell_by_position.get((below_row, cell.column))
        if _is_blank_target(below_cell):
            return _cell_ref(below_row, cell.column)

    return None


def detect_field_labels(sheets: List[SheetInfo]) -> List[FieldLabelCandidate]:
    candidates: List[FieldLabelCandidate] = []

    for sheet in sheets:
        for cell in sheet.cells:
            if not _looks_like_label(cell.value):
                continue

            raw_text = str(cell.value).strip()
            label = _normalize_label_text(raw_text)
            target_cell = _infer_target_cell(sheet, cell)
            confidence = 0.7

            if raw_text.endswith((":", "：")):
                confidence += 0.15

            if cell.is_merged:
                confidence -= 0.05

            candidate = FieldLabelCandidate(
                sheet_name=sheet.sheet_name,
                cell=cell.cell,
                row=cell.row,
                column=cell.column,
                label=label,
                confidence=min(confidence, 0.95),
                reason="匹配字段标签关键词或冒号格式",
            )
            candidate.metadata = {
                "target_cell": target_cell,
                "target_inference": "adjacent_right_then_below",
            }
            candidates.append(candidate)

    return candidates
