from dataclasses import dataclass, field
from typing import Any, Dict, List

from openpyxl.utils import get_column_letter

from app.contracts.template_analysis_result import SheetInfo, TableCandidate
from app.core.table_logic.table_guardrails import (
    is_non_table_layout_row,
    normalize_table_text,
)


@dataclass
class TableRangeCandidate:
    sheet_name: str
    range_ref: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    header_row: int
    data_start_row: int
    data_end_row: int
    confidence: float = 0.0
    reason: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)


def build_table_range_candidate(
    sheet_name: str,
    range_ref: str,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    confidence: float = 0.0,
    reason: str = "",
    metadata: Dict[str, Any] | None = None,
) -> TableRangeCandidate:
    return TableRangeCandidate(
        sheet_name=sheet_name,
        range_ref=range_ref,
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        confidence=confidence,
        reason=reason,
        metadata=dict(metadata or {}),
    )


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_text(cells: list) -> str:
    return " ".join(_text(cell.value) for cell in cells if _text(cell.value))


def _has_explicit_table_title(
    row_index: int,
    cells_by_row: dict[int, list],
) -> bool:
    for previous_row in range(row_index - 1, max(0, row_index - 3), -1):
        previous_text = normalize_table_text(
            _row_text(cells_by_row.get(previous_row, []))
        )
        if "table" in previous_text or "表格" in previous_text:
            return True
    return False


def _row_has_data_in_columns(row_cells: list, min_col: int, max_col: int) -> bool:
    values = [
        cell
        for cell in row_cells
        if min_col <= cell.column <= max_col and _text(cell.value)
    ]
    return len(values) >= 2 and not is_non_table_layout_row(values)


def _data_end_row(
    cells_by_row: dict[int, list],
    header_row: int,
    min_col: int,
    max_col: int,
) -> int:
    data_end = header_row
    last_row = max(cells_by_row.keys(), default=header_row)
    for row_index in range(header_row + 1, last_row + 1):
        row_cells = cells_by_row.get(row_index, [])
        if _row_has_data_in_columns(row_cells, min_col, max_col):
            data_end = row_index
            continue
        break
    return data_end


def detect_tables(sheets: List[SheetInfo]) -> List[TableCandidate]:
    tables: List[TableCandidate] = []

    for sheet in sheets:
        cells_by_row: dict[int, list] = {}
        for cell in sheet.cells:
            if _text(cell.value):
                cells_by_row.setdefault(cell.row, []).append(cell)

        for row_index, row_cells in cells_by_row.items():
            sorted_cells = sorted(row_cells, key=lambda item: item.column)
            text_cells = [cell for cell in sorted_cells if _text(cell.value)]

            if len(text_cells) < 3:
                continue
            if is_non_table_layout_row(text_cells):
                continue

            columns = [cell.column for cell in text_cells]
            if max(columns) - min(columns) + 1 > len(columns) + 2:
                continue

            headers = [_text(cell.value) for cell in text_cells]
            min_col = min(columns)
            max_col = max(columns)
            min_row = row_index
            max_row = _data_end_row(cells_by_row, row_index, min_col, max_col)
            if max_row == row_index and not _has_explicit_table_title(
                row_index,
                cells_by_row,
            ):
                continue
            range_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

            tables.append(
                TableCandidate(
                    sheet_name=sheet.sheet_name,
                    range_ref=range_ref,
                    header_row=row_index,
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    headers=headers,
                    confidence=0.65,
                    reason="row passed Table Detector V1 guardrails with data evidence or explicit table title",
                )
            )

    return tables
